from typing import Literal
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlmodel import Session, select
from io import BytesIO, StringIO
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from server.db import get_session
from server.models.user import User, Recipe
from server.api.deps import check_admin_permission

router = APIRouter()

@router.get("/reports/users/export")
async def export_users(
    format: Literal["csv", "xlsx"] = "csv",
    session: Session = Depends(get_session),
    current_user: dict = Depends(check_admin_permission)
):
    """Export users data to CSV or Excel"""
    
    # Fetch all users
    users = session.exec(select(User)).all()
    
    if format == "csv":
        # Create CSV
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(['ID', 'Name', 'Email', 'Role', 'Provider', 'Created At', 'Disabled'])
        
        # Data
        for user in users:
            writer.writerow([
                user.id,
                user.name or '',
                user.email,
                user.role,
                user.provider,
                user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '',
                'Yes' if user.disabled else 'No'
            ])
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=users_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    else:  # xlsx
        # Create Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Users"
        
        # Header styling
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Header
        headers = ['ID', 'Name', 'Email', 'Role', 'Provider', 'Created At', 'Disabled']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        # Data
        for row, user in enumerate(users, 2):
            ws.cell(row=row, column=1, value=user.id)
            ws.cell(row=row, column=2, value=user.name or '')
            ws.cell(row=row, column=3, value=user.email)
            ws.cell(row=row, column=4, value=user.role)
            ws.cell(row=row, column=5, value=user.provider)
            ws.cell(row=row, column=6, value=user.created_at.strftime('%Y-%m-%d %H:%M:%S') if user.created_at else '')
            ws.cell(row=row, column=7, value='Yes' if user.disabled else 'No')
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=users_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )

@router.get("/reports/recipes/export")
async def export_recipes(
    format: Literal["csv", "xlsx"] = "csv",
    session: Session = Depends(get_session),
    current_user: dict = Depends(check_admin_permission)
):
    """Export recipes data to CSV or Excel"""
    
    # Fetch all recipes
    recipes = session.exec(select(Recipe)).all()
    
    if format == "csv":
        output = StringIO()
        writer = csv.writer(output)
        
        writer.writerow(['ID', 'Title', 'Author', 'Category', 'Difficulty', 'Time', 'Rating', 'Reviews', 'Status', 'Created At'])
        
        for recipe in recipes:
            writer.writerow([
                recipe.id,
                recipe.title,
                recipe.author or '',
                recipe.category,
                recipe.difficulty,
                recipe.time,
                recipe.rating,
                recipe.reviews,
                recipe.status,
                recipe.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(recipe, 'created_at') and recipe.created_at else ''
            ])
        
        output.seek(0)
        return StreamingResponse(
            iter([output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=recipes_{datetime.now().strftime('%Y%m%d')}.csv"}
        )
    
    else:  # xlsx
        wb = Workbook()
        ws = wb.active
        ws.title = "Recipes"
        
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        headers = ['ID', 'Title', 'Author', 'Category', 'Difficulty', 'Time', 'Rating', 'Reviews', 'Status', 'Created At']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for row, recipe in enumerate(recipes, 2):
            ws.cell(row=row, column=1, value=recipe.id)
            ws.cell(row=row, column=2, value=recipe.title)
            ws.cell(row=row, column=3, value=recipe.author or '')
            ws.cell(row=row, column=4, value=recipe.category)
            ws.cell(row=row, column=5, value=recipe.difficulty)
            ws.cell(row=row, column=6, value=recipe.time)
            ws.cell(row=row, column=7, value=recipe.rating)
            ws.cell(row=row, column=8, value=recipe.reviews)
            ws.cell(row=row, column=9, value=recipe.status)
            ws.cell(row=row, column=10, value=recipe.created_at.strftime('%Y-%m-%d %H:%M:%S') if hasattr(recipe, 'created_at') and recipe.created_at else '')
        
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=recipes_{datetime.now().strftime('%Y%m%d')}.xlsx"}
        )
